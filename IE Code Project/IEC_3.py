import json
import sys
import os
import cv2
import easyocr
import traceback
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from bs4 import BeautifulSoup
import pandas as pd
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook




 # Create an empty DataFrame to store the final scraped data
final_df = pd.DataFrame(columns=['IEC Number', 'IEC Issuance Date', 'IEC Status', 'DEL Status', 'IEC Cancelled Date', 'IEC Suspended Date', 'File Number', 'File Date', 'DGFT RA Office', 'Nature of concern/Firm', 'Category of Exporters', 'Address', 'Branch details', 'RCMC Details', 'PAN', 'Total Number Of Branches'])




# Function to handle captcha processing
def solve_captcha_icegate1(browser):
    try:
        iec_input = browser.find_element(By.XPATH, '//*[@id="searchIECode"]')    
        iec_input.send_keys("0388066415")
        time.sleep(3)
            
        captcha_text = process_captcha_icegate1(browser)
        print("captcha_text " , captcha_text)
        #  Input the recognized captcha text
        
        captcha_input = browser.find_element(By.XPATH, '//*[@id="captchaResp"]')
        captcha_input.send_keys(captcha_text)

        time.sleep(3)

        # Click on the element with XPath '//*[@id="pagetable"]/tbody/tr[4]/td[3]/dl/dd/a'
        view_button = browser.find_element(By.XPATH, '//*[@id="pagetable"]/tbody/tr[4]/td[3]/dl/dd/a')
        view_button.click()
        time.sleep(3)

        # Check if the "Invalid Code! Please try again!" message is displayed
        try:
            error_message = browser.find_element(By.XPATH, '//*[@id="pagetable"]/tbody/tr[4]/td[3]/dl/dt[5]/ul/li/span').text
            print("error_message :", error_message) 
            
        
            if "Invalid Code! Please try again!" in error_message:
                return solve_captcha_icegate1(browser)  # Retry solving captcha recursively
                
            else:
                pass
        except Exception as e:
                print("Error in message :", e)
                return True
                
                 
                
    except Exception as e:
        print("Error :", e)
       



# Function to handle captcha processing
def process_captcha_icegate1(browser):
    # Get the captcha image element
    captcha_image = browser.find_element(By.XPATH, '//*[@id="capimg"]')

    # Save the captcha image locally
    with open('captcha.png', 'wb') as file:
        file.write(captcha_image.screenshot_as_png)

    # Load the reader
    reader = easyocr.Reader(['en'])
    result = reader.readtext('captcha.png', detail=0)
    captcha_text = ''.join(result)
    

    return captcha_text

# Disable printing progress information during download
def no_progress(blocknum, bs, size):
    pass

# # Redirect stdout to a file
# sys.stdout = open(os.devnull, 'w')


# Function to handle captcha processing
def solve_captcha_icegate1(browser, iec_code):
    try:
        iec_input = browser.find_element(By.XPATH, '//*[@id="searchIECode"]')    
        iec_input.send_keys(iec_code)
        time.sleep(3)
            
        captcha_text = process_captcha_icegate1(browser)
        print("captcha_text " , captcha_text)
        #  Input the recognized captcha text
        
        captcha_input = browser.find_element(By.XPATH, '//*[@id="captchaResp"]')
        captcha_input.send_keys(captcha_text)

        time.sleep(3)

        # Click on the element with XPath '//*[@id="pagetable"]/tbody/tr[4]/td[3]/dl/dd/a'
        view_button = browser.find_element(By.XPATH, '//*[@id="pagetable"]/tbody/tr[4]/td[3]/dl/dd/a')
        view_button.click()
        time.sleep(3)

        # Check if the "Invalid Code! Please try again!" message is displayed
        try:
            error_message = browser.find_element(By.XPATH, '//*[@id="pagetable"]/tbody/tr[4]/td[3]/dl/dt[5]/ul/li/span').text
            print("error_message :", error_message) 
            
        
            if "Invalid Code! Please try again!" in error_message:
                return solve_captcha_icegate1(browser, iec_code)  # Retry solving captcha recursively
                
            else:
                pass
        except Exception as e:
                print("Error in message :", e)
                return True
                
                 
                
    except Exception as e:
        print("Error :", e)
       



# Function to handle captcha processing
def process_captcha_icegate1(browser):
    # Get the captcha image element
    captcha_image = browser.find_element(By.XPATH, '//*[@id="capimg"]')

    # Save the captcha image locally
    with open('captcha.png', 'wb') as file:
        file.write(captcha_image.screenshot_as_png)

    # Load the reader
    reader = easyocr.Reader(['en'])
    result = reader.readtext('captcha.png', detail=0)
    captcha_text = ''.join(result)
    

    return captcha_text

# Disable printing progress information during download
def no_progress(blocknum, bs, size):
    pass

# # Redirect stdout to a file
# sys.stdout = open(os.devnull, 'w')





def scrape_data_icegate1(browser):
    try:
        
    
        # Locate the element containing the data using its XPath
        data_element = browser.find_element(By.XPATH, '//*[@id="sub_content"]/div[2]')
        
        
        # Get the HTML content of the element
        data_html = data_element.get_attribute('innerHTML')
        print("data_html :", data_html)
        
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(data_html, 'html.parser')
        
                # Find the table
        table = soup.find('table', id='pagetable')
        print("table", table)
        
        
        # Extract the data from the table
        data = {}
        rows = table.find_all('tr')
        address_parts = []
        for row in rows[1:]:  # Skip the header row
            cols = row.find_all('td')
            cols = [col.text.strip() for col in cols]
            if cols:
                field = cols[0].replace('IE Code', 'IEC Number').replace('Name', 'Firm Name')
                value = ' '.join(cols[1:])
                if field in ['Address', '']:  # Combine address parts
                    address_parts.append(value)
                else:
                    data[field] = value

        # Join address parts into a single string
        data['Address'] = ', '.join(address_parts)
        print("data :", data)

        # Create a DataFrame from the extracted data
        df_table1 = pd.DataFrame([data])
        
        
    
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(data_html, 'html.parser')

       
        
        # Find the second table
        tables = soup.find_all('table', id='pagetable')
        if len(tables) >= 2:
            table = tables[1]
            print("Second table:", table)

            # Extract "Total Number Of Branches" value
            total_branches_row = table.find('th', colspan=True)
            if total_branches_row:
                total_branches_value = total_branches_row.text.split(':')[1].strip()
                print("Total Number Of Branches:", total_branches_value)

            # Extract data from the second table
            branch_details = []
            rows = table.find_all('tr')
            current_branch_detail = {}
            for row in rows:
                cols = row.find_all('td')
                cols = [col.text.strip() for col in cols]
                if cols:
                    if cols[0] == 'Branch Serial Number':
                        if current_branch_detail:
                            branch_details.append(current_branch_detail)
                        current_branch_detail = {'Branch Serial Number': cols[1]}
                    elif cols[0] == 'ADDRESS':
                        current_branch_detail['Address'] = ', '.join(cols[1:])

            # Append the last branch detail
            if current_branch_detail:
                branch_details.append(current_branch_detail)

            print("Branch details:", branch_details)

            # Convert "Branch details" to JSON
            branch_details_json = json.dumps(branch_details)
            
            
            
            # Return the scraped data
            return data, total_branches_value, branch_details_json
        
            # Write "Total Number Of Branches" and "Branch details" to Excel
            # df_table2 = pd.DataFrame({'Total Number Of Branches': [total_branches_value], 'Branch details': [branch_details_json]})
            
            # Concatenate df_table1 and df_table2
            # combined_df = pd.concat([df_table1, df_table2], axis=1)

            # # Write the combined DataFrame to Excel
            # combined_df.to_excel('branch_data.xlsx', index=False)
            # print("Data written to 'branch_data.xlsx'")

            # return True

        else:
            print("Second table not found.")
            return data, None, None

    except Exception as e:
        print("Error occurred while scraping data:", e)
        return None, None, None
    



# Replace 'path/to/your/file.xlsx' with the actual path to your Excel file
df = pd.read_excel('IEC_details.xlsx', dtype={'IEC_CODE': str})




def icegate_first(iec_code):
    global scraped_data_df # Add this line to access the global variable
    chrome_options = webdriver.ChromeOptions()
    
    
    # Add preferences to clear cache
    # chrome_prefs = {}
    # chrome_prefs["profile.default_content_settings"] = {"images": 2, "plugins": 2, "popups": 2, "geolocation": 2, 
    #                                                     "notifications": 2, "auto_select_certificate": 2, "fullscreen": 2, 
    #                                                     "mouselock": 2, "mixed_script": 2, "media_stream": 2, 
    #                                                     "media_stream_mic": 2, "media_stream_camera": 2, "protocol_handlers": 2,
    #                                                     "ppapi_broker": 2, "automatic_downloads": 2, "midi_sysex": 2, 
    #                                                     "push_messaging": 2, "ssl_cert_decisions": 2, "metro_switch_to_desktop": 2, 
    #                                                     "protected_media_identifier": 2, "app_banner": 2, "site_engagement": 2, 
    #                                                     "durable_storage": 2}
    # chrome_prefs["profile.default_content_setting_values"] = {"cookies": 2, "images": 2, "javascript": 1, "plugins": 2, "popups": 2, 
    #                                                           "geolocation": 2, "notifications": 2, "auto_select_certificate": 2, 
    #                                                           "fullscreen": 2, "mouselock": 2, "mixed_script": 2, "media_stream": 2, 
    #                                                           "media_stream_mic": 2, "media_stream_camera": 2, "protocol_handlers": 2,
    #                                                           "ppapi_broker": 2, "automatic_downloads": 2, "midi_sysex": 2, 
    #                                                           "push_messaging": 2, "ssl_cert_decisions": 2, "metro_switch_to_desktop": 2, 
    #                                                           "protected_media_identifier": 2, "app_banner": 2, "site_engagement": 2, 
    #                                                           "durable_storage": 2}
    # chrome_options.add_experimental_option("prefs", chrome_prefs)
    
    browser = webdriver.Chrome(options=chrome_options)
    browser.get('https://old.icegate.gov.in/EnqMod/')

    try:
        
   
            
        iec_input = browser.find_element(By.XPATH, '//*[@id="searchIECode"]')    
        iec_input.send_keys(iec_code)
        time.sleep(3)
        
    
    
          # Attempt to solve the captcha
        try:
            captcha_solved_icegate1 = solve_captcha_icegate1(browser, iec_code)
            print("captcha_solved_icegate1:", captcha_solved_icegate1)    
        
        except Exception as e:
            traceback.print_exc()
            print("captcha_solved:", e)    
        
        time.sleep(5)

        if captcha_solved_icegate1:
        # Wait for a few seconds before checking for the presence of the message
            time.sleep(3)
            
             # Check if the message "Details for this IEC Number is not available." exists
            try:
                unavailable_message = browser.find_element(By.XPATH, '//*[@id="sub_content"]/div[2]').text
                if "No Record Found" in unavailable_message:
                    # icegate_second()  # Call the icegate function if the message exists
                    pass
                else:
                    # Proceed with scraping the data if the message doesn't exist
                    # Define the column names to be set as NULL
                    null_columns = [
                        "IEC Issuance Date", "IEC Status", "DEL Status", "IEC Cancelled Date",
                        "IEC Suspended Date", "File Number", "File Date", "DGFT RA Office", "Category of Exporters",
                        "RCMC DETAILS", "Nature of concern/Firm"
                    ]
                    
                    # Create a dictionary with NULL values for the specified columns
                    null_data = {column: ["NULL"] for column in null_columns}
                    
                    
                    scraped_data = scrape_data_icegate1(browser)
                    if scraped_data is not None:
                        data, total_branches_value, branch_details_json = scraped_data
                        # Create a new row in the DataFrame with the scraped data
                        new_row = {
                            'IEC Issuance Date': null_data.get('IEC Issuance Date',''),
                            'DEL Status': null_data.get('DEL Status',''),
                            'IEC Number': data.get('IEC Number', ''),
                            'Firm Name': data.get('Firm Name', ''),
                            'Address': data.get('Address', ''),
                            'PAN' : data.get('PAN', ''),
                            'IEC Status' : data.get('IEC Status', ''),
                            'Total Number Of Branches': total_branches_value,
                            'Branch details': branch_details_json
                        }
                        
                         # Create a new DataFrame to store scraped data
                        # scraped_data_df = pd.DataFrame(columns=['IEC Number' ,'IEC Issuance Date','IEC Status', 'DEL Status', 'IEC Cancelled Date', 'IEC Suspended Date', 'File Number','File Date','DGFT RA Office',
                        #                 'Nature of concern/Firm', 'Category of Exporters' , 'Address', 'Branch details', 'RCMC Details' ,'PAN', 'Total Number Of Branches'])
                        
                        # Create a DataFrame from the row data
                        # icegate_data_df = pd.DataFrame([new_row])
                        
                        # final_df = pd.concat([final_df, icegate_data_df], ignore_index=True)

                        # Return the scraped data DataFrame
                        # return icegate_data_df
                        scraped_data_df= pd.concat([scraped_data_df, pd.DataFrame([new_row])], ignore_index=True)
                        
                        # Write the scraped data to an Excel file
                        # scraped_data_df.to_excel('scraped_data.xlsx', index=False)
                   
                        
            except NoSuchElementException:
                # If the element is not found, proceed with scraping the data
                # scrape_data_icegate1(browser)
                pass
            
    except Exception as e:
        traceback.print_exc()
        print("wrong xpath", e)

    finally:
        browser.quit()

        # Restore stdout
        sys.stdout = sys.__stdout__






# Function to handle captcha processing
def solve_captcha_dgft(browser):
    captcha_text = process_captcha_dgft(browser)
    print("captcha_text " , captcha_text)
    
    captcha_input = browser.find_element(By.XPATH, '//*[@id="txt_Captcha"]')
    captcha_input.clear()
    captcha_input.send_keys(captcha_text)

    time.sleep(3)

    view_button = browser.find_element(By.XPATH, '//*[@id="viewIEC1"]')
    view_button.click()
    time.sleep(3)

     # Check if the "Please enter valid captcha code" message is displayed
    error_message = browser.find_element(By.XPATH, '//*[@id="incCaptcha"]').text
    print(" error_message",error_message) 
    
    if "Please enter valid captcha code" in error_message:
        return solve_captcha_dgft(browser)  # Retry solving captcha recursively
        
    else:
        return True


#Function to handle captcha processing
def process_captcha_dgft(browser):
    # Get the captcha image element
    captcha_image = browser.find_element(By.XPATH, '//*[@id="captcha"]')


    # Save the captcha image locally
    with open('captcha.png', 'wb') as file:
        file.write(captcha_image.screenshot_as_png)
        
    # Preprocess the captcha image
    captcha_image = cv2.imread('captcha.png')
    gray_image = cv2.cvtColor(captcha_image, cv2.COLOR_BGR2GRAY)
    # _, binary_image = cv2.threshold(gray_image, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)
    # blurred_image = cv2.GaussianBlur(binary_image, (3, 3), 0)
    cv2.imwrite('preprocessed_captcha.png', gray_image)
    
    # Load the reader
    reader = easyocr.Reader(['en'])
    result = reader.readtext('captcha.png', detail=0)
    captcha_text = ''.join(result)
    print("captcha_text", captcha_text)

    return captcha_text




# Create a new DataFrame to store scraped data
scraped_data_df = pd.DataFrame(columns=['IEC Number' ,'IEC Issuance Date','IEC Status', 'DEL Status', 'IEC Cancelled Date', 'IEC Suspended Date', 'File Number','File Date','DGFT RA Office',
                                        'Nature of concern/Firm', 'Category of Exporters' ,'Firm Name', 'Address', 'Branch details', 'RCMC Details' ,'PAN', 'Total Number Of Branches'])


def scrape_data_dgft(browser):
    global scraped_data_df  # Access the global DataFrame
    try:
        
       
        symbol_click =  browser.find_element(By.XPATH, '//*[@id="custom-accordion"]/div[2]/div[1]/a')
        
        symbol_click.click
        
         # Locate the element containing the data using its XPath
        rcmc_element = browser.find_element(By.XPATH, '//*[@id="rcmc"]/div')
        print("rcmc_element :" , rcmc_element)
        
         # Get the HTML content of the element
        rcmc_html = rcmc_element.get_attribute('innerHTML')
        
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(rcmc_html, 'html.parser')
        
        
        rcmc_branch_details = [] 
        
        # Scraping data from table format
        table = soup.find('table', class_='table table-hover custom-datatable')
        print("table :" , table)
        if table:
            headers = [th.text.strip() for th in table.find_all('th')]
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all('td')
                if cells:
                    # Extract text from cells and add to all_branch_details
                    row_details = [cell.text.strip() for cell in cells]
                    rcmc_branch_details.append(dict(zip(headers, row_details)))
        
        # Convert all_branch_details to JSON string
        rcmc_details_json = json.dumps(rcmc_branch_details)

        # Create a Pandas DataFrame with the JSON data
        rcmc_df = pd.DataFrame({"RCMC DETAILS": [rcmc_details_json]})
        print("rcmc_df", rcmc_df)
        
        
        # Locate the element containing the data using its XPath
        data_element = browser.find_element(By.XPATH, '//*[@id="iecdetails"]')
        
        # Get the HTML content of the element
        data_html = data_element.get_attribute('innerHTML')
        
        # Parse the HTML content using BeautifulSoup
        soup = BeautifulSoup(data_html, 'html.parser')

        # Scraping data from div elements with multiple rows
        div_elements = soup.find_all('div', class_='card-body')
                      
        div_details = []
        for div_element in div_elements:
            rows = div_element.find_all('div', class_='form-group')
            details = {}
            for row in rows:
                label_element = row.find('label', class_='font-12 font-weight-semi-bold')
                if label_element:
                    label = label_element.text.strip()
                    value_element = row.find('p', class_='font-12 text-gray')
                    if value_element:
                        value = value_element.text.strip()
                        details[label] = value
            div_details.append(details)
            print("div_details    :" , div_details)
            print("details    :" , details)

        all_branch_details = []  # Initialize a list to store all branch details
        
        # Loop through each pagination link until there is no "Next" button
        while True:
            # Locate the element containing the data using its XPath
            data_element = browser.find_element(By.XPATH, '//*[@id="iecdetails"]')
            
            # Get the HTML content of the element
            data_html = data_element.get_attribute('innerHTML')
            
            # Parse the HTML content using BeautifulSoup
            soup = BeautifulSoup(data_html, 'html.parser')

            # Scraping data from table format
            table = soup.find('table', class_='table table-hover custom-datatable dataTable no-footer')
            if table:
                headers = [th.text.strip() for th in table.find_all('th')]
                rows = table.find_all('tr')
                for row in rows:
                    cells = row.find_all('td')
                    if cells:
                        # Extract text from cells and add to all_branch_details
                        row_details = [cell.text.strip() for cell in cells]
                        all_branch_details.append(dict(zip(headers, row_details)))
            
            # Click the "Next" button to navigate to the next page
            next_button = browser.find_element(By.ID, 'branchTable_next')
            if 'disabled' in next_button.get_attribute('class'):
                # If the "Next" button is disabled, break the loop
                break
            else:
                next_button.click()
                
                # Wait for the table to load (you may need to implement this)
                # Add code to wait for the table to load here
                time.sleep(2)  # Adjust the delay according to your page loading time
        
        # Convert all_branch_details to JSON string
        branch_details_json = json.dumps(all_branch_details)
         
         # Return the scraped data
        print("RCMC=========="  , rcmc_details_json, details, branch_details_json)
        
        
         # Define the column names to be set as NULL
        null_columns = ["PAN", "Total Number Of Branches"]
        
        # Create a dictionary with NULL values for the specified columns
        null_data = {column: ["NULL"] for column in null_columns}
         # Create a new row in the DataFrame with the scraped data
        new_row = {
            'IEC Number' : details.get('IEC Number', ''),
            'IEC Issuance Date': details.get('IEC Issuance Date',''),
            'IEC Status' : details.get('IEC Status',''),
            'DEL Status': details.get('DEL Status',''),
            'IEC Cancelled Date' : details.get('IEC Cancelled Date',''),
            'IEC Suspended Date': details.get('IEC Suspended Date',''),
            'File Number': details.get('File Number',''),
            'File Date': details.get('File Date',''),
            'DGFT RA Office': details.get('DGFT RA Office',''),
            'Nature of concern/Firm': details.get('Nature of concern/Firm',''),
            'Category of Exporters' : details.get('Category of Exporters',''),
            'Firm Name' : details.get('Firm Name',''),
            'Address' : details.get('Address',''),
            'Branch details': branch_details_json,
            'RCMC Details' : rcmc_details_json,
            'PAN' : null_data.get('PAN',''),
            'Total Number Of Branches': null_data.get('Total Number Of Branches','')
            
        }
        
    
        scraped_data_df = pd.concat([scraped_data_df, pd.DataFrame([new_row])], ignore_index=True)
        
        # return True
        # scraped_data_df.to_excel('scraped_data.xlsx', index=False)
        # return scraped_data_df 
        # print("scraped_data_df ", scraped_data_df)
        
        
        
        # return rcmc_details_json, details, branch_details_json
        # # Define the path to save the Excel file
        # excel_file = "scraped_data.xlsx"

        # # Create a Pandas DataFrame with the JSON data
        # branch_df = pd.DataFrame({"BRANCH DETAILS": [branch_details_json]})
        
        # # Write the DataFrame to an Excel file
        # branch_df.to_excel(excel_file, index=False)
        
        # print(f"Data has been exported to {excel_file}")
        
        # # Convert scraped data to DataFrames
        # table_df = pd.DataFrame(all_branch_details)
        # div_df = pd.DataFrame(div_details)
        
        # # Remove \n\t from DataFrame columns
        # table_df = table_df.replace(r'\n\t','', regex=True)
        # div_df = div_df.replace(r'\n\t','', regex=True)
        # branch_df = branch_df.replace(r'\n\t','', regex=True)
        # rcmc_df = rcmc_df.replace(r'\n\t','', regex=True)
        
        
        
        # # Define columns to set as NULL
        # null_columns = ["PAN", "Total Number Of Branches"]

        # # Set NULL values for specified columns
        # for column in null_columns:
        #     branch_df[column] = "NULL"
        
        

        # # Merge all DataFrames and write to Excel
        # merged_df = pd.concat([pd.DataFrame({"IEC NUMBER": ["0301014175"]}), div_df, branch_df , rcmc_df], axis=1)
        # # with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
        # #     merged_df.to_excel(writer, sheet_name='Merged Data', index=False)
            
            
        # # Check if the Excel file exists
        # if os.path.isfile(excel_file):
        #     # Load the existing Excel file
        #     book = load_workbook(excel_file)
        #     writer = pd.ExcelWriter(excel_file, engine='openpyxl', mode='a')

        #     # Check if the 'Merged Data' sheet exists
        #     if 'Merged Data' in book.sheetnames:
        #         # Overwrite the existing 'Merged Data' sheet
        #         with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        #             merged_df.to_excel(writer, sheet_name='Merged Data', index=False, startrow=0)
        #     else:
        #         # Create a new 'Merged Data' sheet
        #         with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
        #             merged_df.to_excel(writer, sheet_name='Merged Data', index=False)
        # else:
        #     # Create the Excel file and write data to it
        #     merged_df.to_excel(excel_file, sheet_name='Merged Data', index=False)

        # print("merged_df :", merged_df)

        # print(f"Data has been exported to {excel_file}")
            
        # return True

    except Exception as e:
        print("Error occurred while scraping data:", e)
        return False
    




    
# Disable printing progress information during download
def no_progress(blocknum, bs, size):
    pass




# Replace 'path/to/your/file.xlsx' with the actual path to your Excel file
df = pd.read_excel('IEC_details.xlsx', dtype={'IEC_CODE': str})










# Redirect stdout to a file
# sys.stdout = open(os.devnull, 'w')
def dgft(iec_code, firm_name):
    chrome_options = webdriver.ChromeOptions()
    browser = webdriver.Chrome(options=chrome_options)
    browser.get('https://www.dgft.gov.in/CP/?opt=view-any-ice')

    try:
        next_xpath = browser.find_element(By.XPATH, '//*[@id="mainSectionWrap"]/div[3]/div/div[2]/div[1]')
        next_xpath.click()
        time.sleep(2)

        iec_input = browser.find_element(By.XPATH, '//*[@id="iecNo"]')
        entity_input = browser.find_element(By.XPATH, '//*[@id="entity"]')
        iec_input.send_keys(iec_code)
        time.sleep(3)
        entity_input.send_keys(firm_name)
        time.sleep(3)

        
        # Attempt to solve the captcha
        try:
            captcha_solved_dgft = solve_captcha_dgft(browser)
            print("captcha_solved:", captcha_solved_dgft)    
        
        except Exception as e:
            traceback.print_exc()
            print("captcha_solved:", e)    
        
        time.sleep(5)
        
        if captcha_solved_dgft:
            # Wait for a few seconds before checking for the presence of the message
            time.sleep(3)
            
            # Check if the message "Details for this IEC Number is not available." exists
            try:
                unavailable_message = browser.find_element(By.XPATH, '/html/body/div[16]/div/div/div/div[1]').text
                if "Details for this IEC Number is not available." in unavailable_message:
                    icegate_first(iec_code)  # Call the icegate function if the message exists
                    # pass
                    
            
                else:
                     # Proceed with scraping the data if the message doesn't exist
                     # Define the column names to be set as NULL
                     null_columns = ["PAN", "Total Number Of Branches"]
                        
                     # Create a dictionary with NULL values for the specified columns
                     null_data = {column: ["NULL"] for column in null_columns}
                    
                    
                      # Proceed with scraping the data if the message doesn't exist
                     scraped_data = scrape_data_dgft(browser)
                     print("scraped_dat====", scraped_data)
                     if scraped_data is not None:
                        rcmc_details_json, details, branch_details_json = scraped_data
                        # Create a new row in the DataFrame with the scraped data
                        new_row = {
                            'IEC Number' : details.get('IEC Number', ''),
                            'IEC Issuance Date': details.get('IEC Issuance Date',''),
                            'IEC Status' : details.get('IEC Status',''),
                            'DEL Status': details.get('DEL Status',''),
                            'IEC Cancelled Date' : details.get('IEC Cancelled Date',''),
                            'IEC Suspended Date': details.get('IEC Suspended Date',''),
                            'File Number': details.get('File Number',''),
                            'File Date': details.get('File Date',''),
                            'DGFT RA Office': details.get('DGFT RA Office',''),
                            'Nature of concern/Firm': details.get('Nature of concern/Firm',''),
                            'Category of Exporters' : details.get('Category of Exporters',''),
                            'Firm Name' : details.get('Firm Name',''),
                            'Address' : details.get('Address',''),
                            'Branch details': branch_details_json,
                            'RCMC Details' : rcmc_details_json,
                            'PAN' : null_data.get('PAN',''),
                            'Total Number Of Branches': null_data.get('Total Number Of Branches','')
                            
                        }
                        
                        scraped_data_df= pd.concat([scraped_data_df, pd.DataFrame([new_row])], ignore_index=True)
            except NoSuchElementException:
                # If the element is not found, proceed with scraping the data
                scrape_data_dgft(browser)

    except Exception as e:
        traceback.print_exc()
        print("An error occurred:", e)

    finally:
        browser.quit()
    
    

def read_excel():
    try:
        for index, row in df.head(2).iterrows(): 
            iec_code = row['IEC_CODE']
            print("iec_code", iec_code)
            firm_name = row['FIRM NAME']
            
            # Call your function or code with the IEC code and firm name
            dgft(iec_code, firm_name)
        
        # Write the scraped data to an Excel file after processing all rows
        scraped_data_df.to_excel('scraped_data.xlsx', index=False)
        
    except Exception as e:
        traceback.print_exc()
        print("Excel reading  error", e)
        
    #  # Write the scraped data to an Excel file   
    # scraped_data_df.to_excel('scraped_data.xlsx', index=False)
     
    
read_excel()

